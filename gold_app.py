import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import sys

class GoldTransactionApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Gold Transaction Manager - مدیریت معاملات طلا")
        self.root.geometry("700x600")
        self.root.configure(bg='#f0f0f0')
        
        # Set window icon (optional)
        try:
            self.root.iconbitmap(default='icon.ico')
        except:
            pass
        
        # Center the window on screen
        self.center_window()
        
        # Constants
        self.FILE_NAME = 'transactions.xlsx'
        self.METHQAL_TO_GRAM = 4.3317
        
        # Initialize Excel file
        self.init_file()
        
        # Create GUI
        self.create_widgets()
        
    def center_window(self):
        """Center the window on screen"""
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        pos_x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        pos_y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{pos_x}+{pos_y}')
        
    def init_file(self):
        """Create Excel file if it doesn't exist"""
        try:
            if not os.path.exists(self.FILE_NAME):
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Transactions"
                headers = ['type', 'date', 'weight', 'karat', 'price_per_gram', 'price_per_methqal', 
                          'wage', 'profit', 'tax', 'total_price', 'note', 'total_gold']
                ws.append(headers)
                wb.save(self.FILE_NAME)
                messagebox.showinfo("Information", f"Excel file created successfully!\nFile location: {os.path.abspath(self.FILE_NAME)}")
        except Exception as e:
            messagebox.showerror("Error", f"Could not create Excel file: {str(e)}")
    
    def create_widgets(self):
        # Main title
        title_label = tk.Label(self.root, text="مدیریت معاملات طلا", 
                              font=("Arial", 16, "bold"), bg='#f0f0f0')
        title_label.pack(pady=20)
        
        # Button frame
        button_frame = tk.Frame(self.root, bg='#f0f0f0')
        button_frame.pack(pady=20)
        
        # Buy button
        buy_button = tk.Button(button_frame, text="ثبت خرید طلا", 
                              command=self.open_buy_dialog,
                              bg='#4CAF50', fg='white', 
                              font=("Arial", 12, "bold"),
                              width=20, height=2)
        buy_button.pack(pady=10)
        
        # Sell button
        sell_button = tk.Button(button_frame, text="ثبت فروش طلا", 
                               command=self.open_sell_dialog,
                               bg='#f44336', fg='white',
                               font=("Arial", 12, "bold"),
                               width=20, height=2)
        sell_button.pack(pady=10)
        
        # Inventory button
        inventory_button = tk.Button(button_frame, text="نمایش موجودی", 
                                    command=self.show_inventory,
                                    bg='#2196F3', fg='white',
                                    font=("Arial", 12, "bold"),
                                    width=20, height=2)
        inventory_button.pack(pady=10)
        
        # View transactions button
        view_button = tk.Button(button_frame, text="نمایش تمام تراکنش‌ها", 
                               command=self.show_all_transactions,
                               bg='#FF9800', fg='white',
                               font=("Arial", 12, "bold"),
                               width=20, height=2)
        view_button.pack(pady=10)
        
        # Exit button
        exit_button = tk.Button(button_frame, text="خروج", 
                               command=self.root.quit,
                               bg='#607D8B', fg='white',
                               font=("Arial", 12, "bold"),
                               width=20, height=2)
        exit_button.pack(pady=10)
        
        # Status label
        self.status_label = tk.Label(self.root, text="Developed by masoud89", 
                                    font=("Arial", 10), bg='#f0f0f0')
        self.status_label.pack(side=tk.BOTTOM, pady=10)
    
    def calculate_gold_value(self, weight, karat):
        """Calculate gold value using the formula: weight * (karat / 750)"""
        return weight * (karat / 750)
    
    def calculate_total_price_buy(self, price_per_methqal, karat, weight, wage):
        """Calculate total price for buy transaction"""
        base_price = (price_per_methqal / self.METHQAL_TO_GRAM) * (karat / 750) * weight
        total_price = base_price + (base_price * wage / 100)
        return round(total_price, 2)
    
    def calculate_total_price_sell(self, price_per_gram, karat, weight, wage, profit, tax):
        """Calculate total price for sell transaction"""
        base_price = price_per_gram * (karat / 750) * weight
        wage_amount = base_price * wage / 100
        base_with_wage = base_price + wage_amount
        profit_amount = base_with_wage * profit / 100
        tax_amount = profit_amount * tax / 100
        total_price = base_price + wage_amount + profit_amount + tax_amount
        return round(total_price, 2)
    
    def get_last_total_gold(self):
        """Get the total_gold from the last transaction"""
        transactions = self.load_transactions()
        if not transactions:
            return 0.0
        return float(transactions[-1]['total_gold']) if transactions[-1]['total_gold'] is not None else 0.0
    
    def open_buy_dialog(self):
        """Open dialog for buying gold"""
        dialog = tk.Toplevel(self.root)
        dialog.title("ثبت خرید طلا")
        dialog.geometry("400x450")
        dialog.configure(bg='#f0f0f0')
        
        # Weight input
        tk.Label(dialog, text="وزن خرید (گرم):", font=("Arial", 10), bg='#f0f0f0').pack(pady=5)
        weight_entry = tk.Entry(dialog, font=("Arial", 10))
        weight_entry.pack(pady=5)
        
        # Karat input
        tk.Label(dialog, text="عیار طلا (مثلاً 750 یا 900):", font=("Arial", 10), bg='#f0f0f0').pack(pady=5)
        karat_entry = tk.Entry(dialog, font=("Arial", 10))
        karat_entry.pack(pady=5)
        
        # Price per methqal input
        tk.Label(dialog, text="قیمت هر مثقال (تومان):", font=("Arial", 10), bg='#f0f0f0').pack(pady=5)
        price_entry = tk.Entry(dialog, font=("Arial", 10))
        price_entry.pack(pady=5)
        
        # Wage input
        tk.Label(dialog, text="اجرت (درصد):", font=("Arial", 10), bg='#f0f0f0').pack(pady=5)
        wage_entry = tk.Entry(dialog, font=("Arial", 10))
        wage_entry.pack(pady=5)
        
        # Note input
        tk.Label(dialog, text="توضیح (اختیاری):", font=("Arial", 10), bg='#f0f0f0').pack(pady=5)
        note_entry = tk.Entry(dialog, font=("Arial", 10))
        note_entry.pack(pady=5)
        
        # Save button
        def save_buy():
            try:
                weight = float(weight_entry.get())
                karat = int(karat_entry.get())
                price_per_methqal = float(price_entry.get())
                wage = float(wage_entry.get()) if wage_entry.get() else 0.0
                price_per_gram = round(price_per_methqal / self.METHQAL_TO_GRAM, 2)
                note = note_entry.get()
                date = datetime.now().strftime("%Y-%m-%d")
                
                # Default values for buy transaction
                profit = 0.0
                tax = 0.0
                
                # Calculate total price
                total_price = self.calculate_total_price_buy(price_per_methqal, karat, weight, wage)
                
                # Calculate new total_gold
                last_total = self.get_last_total_gold()
                gold_value = self.calculate_gold_value(weight, karat)
                new_total_gold = last_total + gold_value
                
                self.save_transaction('buy', date, weight, karat, price_per_gram, price_per_methqal, 
                                    wage, profit, tax, total_price, note, new_total_gold)
                
                messagebox.showinfo("موفق", f"خرید ثبت شد.\nقیمت هر گرم: {price_per_gram:,.0f} تومان\nقیمت کل: {total_price:,.0f} تومان\nموجودی جدید: {new_total_gold:.4f} گرم طلای خالص")
                dialog.destroy()
                self.status_label.config(text="خرید جدید ثبت شد")
            except ValueError:
                messagebox.showerror("خطا", "لطفاً مقادیر صحیح وارد کنید")
        
        tk.Button(dialog, text="ثبت خرید", command=save_buy, 
                 bg='#4CAF50', fg='white', font=("Arial", 10, "bold")).pack(pady=20)
    
    def open_sell_dialog(self):
        """Open dialog for selling gold"""
        dialog = tk.Toplevel(self.root)
        dialog.title("ثبت فروش طلا")
        dialog.geometry("400x550")
        dialog.configure(bg='#f0f0f0')
        
        # Weight input
        tk.Label(dialog, text="وزن فروش (گرم):", font=("Arial", 10), bg='#f0f0f0').pack(pady=5)
        weight_entry = tk.Entry(dialog, font=("Arial", 10))
        weight_entry.pack(pady=5)
        
        # Karat input
        tk.Label(dialog, text="عیار طلا (مثلاً 750 یا 900):", font=("Arial", 10), bg='#f0f0f0').pack(pady=5)
        karat_entry = tk.Entry(dialog, font=("Arial", 10))
        karat_entry.pack(pady=5)
        
        # Price per gram input
        tk.Label(dialog, text="قیمت فروش هر مثقال (تومان):", font=("Arial", 10), bg='#f0f0f0').pack(pady=5)
        price_entry = tk.Entry(dialog, font=("Arial", 10))
        price_entry.pack(pady=5)
        
        # Wage input
        tk.Label(dialog, text="اجرت (درصد):", font=("Arial", 10), bg='#f0f0f0').pack(pady=5)
        wage_entry = tk.Entry(dialog, font=("Arial", 10))
        wage_entry.pack(pady=5)
        
        # Profit input
        tk.Label(dialog, text="سود (درصد):", font=("Arial", 10), bg='#f0f0f0').pack(pady=5)
        profit_entry = tk.Entry(dialog, font=("Arial", 10))
        profit_entry.pack(pady=5)
        
        # Tax input
        tk.Label(dialog, text="مالیات (درصد):", font=("Arial", 10), bg='#f0f0f0').pack(pady=5)
        tax_entry = tk.Entry(dialog, font=("Arial", 10))
        tax_entry.pack(pady=5)
        
        # Note input
        tk.Label(dialog, text="توضیح (اختیاری):", font=("Arial", 10), bg='#f0f0f0').pack(pady=5)
        note_entry = tk.Entry(dialog, font=("Arial", 10))
        note_entry.pack(pady=5)
        
        # Save button
        def save_sell():
            try:
                weight = float(weight_entry.get())
                karat = int(karat_entry.get())
                price_per_methqal = float(price_entry.get())
                price_per_gram = round(price_per_methqal / self.METHQAL_TO_GRAM, 2)
                wage = float(wage_entry.get()) if wage_entry.get() else 0.0
                profit = float(profit_entry.get()) if profit_entry.get() else 0.0
                tax = float(tax_entry.get()) if tax_entry.get() else 0.0
                note = note_entry.get()
                date = datetime.now().strftime("%Y-%m-%d")
                
                # Calculate total price
                total_price = self.calculate_total_price_sell(price_per_gram, karat, weight, wage, profit, tax)
                
                # Calculate new total_gold
                last_total = self.get_last_total_gold()
                gold_value = self.calculate_gold_value(weight, karat)
                new_total_gold = last_total - gold_value
                
                # Check if selling more than available
                if new_total_gold < 0:
                    messagebox.showwarning("هشدار", f"موجودی فعلی: {last_total:.4f} گرم طلای خالص\nمقدار فروش: {gold_value:.4f} گرم طلای خالص\nموجودی منفی خواهد شد!")
                
                self.save_transaction('sell', date, weight, karat, price_per_gram, price_per_methqal, 
                                    wage, profit, tax, total_price, note, new_total_gold)
                
                messagebox.showinfo("موفق", f"فروش ثبت شد\nقیمت کل: {total_price:,.0f} تومان\nموجودی جدید: {new_total_gold:.4f} گرم طلای خالص")
                dialog.destroy()
                self.status_label.config(text="فروش جدید ثبت شد")
            except ValueError:
                messagebox.showerror("خطا", "لطفاً مقادیر صحیح وارد کنید")
        
        tk.Button(dialog, text="ثبت فروش", command=save_sell, 
                 bg='#f44336', fg='white', font=("Arial", 10, "bold")).pack(pady=20)
    
    def show_inventory(self):
        """Show current inventory based on latest record's total_gold"""
        transactions = self.load_transactions()
        
        if not transactions:
            messagebox.showinfo("موجودی", "هیچ تراکنشی یافت نشد")
            return
        
        # Get total_gold from the latest record
        latest_total_gold = float(transactions[-1]['total_gold']) if transactions[-1]['total_gold'] is not None else 0.0
        
        messagebox.showinfo("موجودی", f"موجودی فعلی طلای خالص: {latest_total_gold:.4f} گرم")
        self.status_label.config(text=f"موجودی: {latest_total_gold:.4f} گرم طلای خالص")
    
    def show_all_transactions(self):
        """Show all transactions in a new window"""
        transactions = self.load_transactions()
        
        if not transactions:
            messagebox.showinfo("اطلاع", "هیچ تراکنشی یافت نشد")
            return
        
        # Create new window
        trans_window = tk.Toplevel(self.root)
        trans_window.title("تمام تراکنش‌ها")
        trans_window.geometry("1200x500")
        
        # Create frame for treeview and scrollbar
        tree_frame = tk.Frame(trans_window)
        tree_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Create treeview
        tree = ttk.Treeview(tree_frame, columns=('type', 'date', 'weight', 'karat', 'price_gram', 'price_methqal', 'wage', 'profit', 'tax', 'total_price', 'note', 'total_gold'), show='headings')
        
        # Define headings
        tree.heading('type', text='نوع')
        tree.heading('date', text='تاریخ')
        tree.heading('weight', text='وزن')
        tree.heading('karat', text='عیار')
        tree.heading('price_gram', text='قیمت/گرم')
        tree.heading('price_methqal', text='قیمت/مثقال')
        tree.heading('wage', text='اجرت%')
        tree.heading('profit', text='سود%')
        tree.heading('tax', text='مالیات%')
        tree.heading('total_price', text='قیمت کل')
        tree.heading('note', text='توضیح')
        tree.heading('total_gold', text='موجودی کل')
        
        # Set column widths
        tree.column('type', width=60)
        tree.column('date', width=80)
        tree.column('weight', width=70)
        tree.column('karat', width=60)
        tree.column('price_gram', width=80)
        tree.column('price_methqal', width=90)
        tree.column('wage', width=60)
        tree.column('profit', width=60)
        tree.column('tax', width=60)
        tree.column('total_price', width=90)
        tree.column('note', width=120)
        tree.column('total_gold', width=90)
        
        # Insert data
        for trans in transactions:
            tree.insert('', 'end', values=(
                'خرید' if trans['type'] == 'buy' else 'فروش',
                trans['date'],
                trans['weight'],
                trans['karat'],
                trans['price_per_gram'],
                trans['price_per_methqal'],
                trans.get('wage', 0),
                trans.get('profit', 0),
                trans.get('tax', 0),
                f"{float(trans.get('total_price', 0)):,.0f}" if trans.get('total_price') else "0",
                trans['note'],
                f"{float(trans['total_gold']):.4f}" if trans['total_gold'] is not None else "0.0000"
            ))
        
        # Scrollbars
        v_scrollbar = ttk.Scrollbar(tree_frame, orient='vertical', command=tree.yview)
        h_scrollbar = ttk.Scrollbar(tree_frame, orient='horizontal', command=tree.xview)
        tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        # Pack treeview and scrollbars
        tree.grid(row=0, column=0, sticky='nsew')
        v_scrollbar.grid(row=0, column=1, sticky='ns')
        h_scrollbar.grid(row=1, column=0, sticky='ew')
        
        # Configure grid weights
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
    
    def save_transaction(self, t_type, date, weight, karat, price_per_gram, price_per_methqal, wage, profit, tax, total_price, note, total_gold):
        """Save transaction to Excel file"""
        wb = openpyxl.load_workbook(self.FILE_NAME)
        ws = wb["Transactions"]
        ws.append([t_type, date, weight, karat, price_per_gram, price_per_methqal, wage, profit, tax, total_price, note, total_gold])
        wb.save(self.FILE_NAME)
    
    def load_transactions(self):
        """Load all transactions from Excel file"""
        if not os.path.exists(self.FILE_NAME):
            return []
        
        wb = openpyxl.load_workbook(self.FILE_NAME)
        ws = wb["Transactions"]
        transactions = []
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            
            # Handle both old and new format
            if len(row) >= 12:  # New format with all fields
                transactions.append({
                    'type': row[0],
                    'date': row[1],
                    'weight': row[2],
                    'karat': row[3],
                    'price_per_gram': row[4],
                    'price_per_methqal': row[5],
                    'wage': row[6] if row[6] is not None else 0,
                    'profit': row[7] if row[7] is not None else 0,
                    'tax': row[8] if row[8] is not None else 0,
                    'total_price': row[9] if row[9] is not None else 0,
                    'note': row[10],
                    'total_gold': row[11] if row[11] is not None else 0.0
                })
            else:  # Old format - backward compatibility
                transactions.append({
                    'type': row[0],
                    'date': row[1],
                    'weight': row[2],
                    'karat': row[3],
                    'price_per_gram': row[4],
                    'price_per_methqal': row[5],
                    'wage': 0,
                    'profit': 0,
                    'tax': 0,
                    'total_price': 0,
                    'note': row[6] if len(row) > 6 else '',
                    'total_gold': row[7] if len(row) > 7 and row[7] is not None else 0.0
                })
        
        return transactions

if __name__ == '__main__':
    root = tk.Tk()
    app = GoldTransactionApp(root)
    root.mainloop()
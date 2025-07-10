import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

FILE_NAME = 'transactions.xlsx'
METHQAL_TO_GRAM = 4.3317

# -------------------------------------------------
# ساخت فایل اکسل در صورت عدم وجود
def init_file():
    if not os.path.exists(FILE_NAME):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transactions"
        headers = ['type', 'date', 'weight', 'karat', 'price_per_gram', 'price_per_methqal', 'note']
        ws.append(headers)
        wb.save(FILE_NAME)

# -------------------------------------------------
# ذخیره تراکنش جدید در اکسل
def save_transaction(t_type, date, weight, karat,
                     price_per_gram, price_per_methqal, note):
    wb = openpyxl.load_workbook(FILE_NAME)
    ws = wb["Transactions"]
    ws.append([t_type, date, weight, karat, price_per_gram, price_per_methqal, note])
    wb.save(FILE_NAME)

# -------------------------------------------------
# خواندن تمام تراکنش‌ها از اکسل
def load_transactions():
    if not os.path.exists(FILE_NAME):
        return []
    wb = openpyxl.load_workbook(FILE_NAME)
    ws = wb["Transactions"]
    transactions = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        transactions.append({
            'type': row[0],
            'date': row[1],
            'weight': row[2],
            'karat': row[3],
            'price_per_gram': row[4],
            'price_per_methqal': row[5],
            'note': row[6]
        })
    return transactions

# -------------------------------------------------
# محاسبه و نمایش موجودی
def show_inventory():
    total = 0.0
    for t in load_transactions():
        if t['type'] == 'buy':
            total += float(t['weight'])
        elif t['type'] == 'sell':
            total -= float(t['weight'])
    print(f"\n📦 موجودی فعلی طلا: {total:.2f} گرم")

# -------------------------------------------------
def main():
    init_file()
    while True:
        print("\n--- منوی اصلی ---")
        print("1. ثبت خرید طلا")
        print("2. ثبت فروش طلا")
        print("3. نمایش موجودی")
        print("4. خروج")
        choice = input("انتخاب: ")

        if choice == '1':
            try:
                weight = float(input("🔸 وزن خرید (گرم): "))
                karat = int(input("🔸 عیار طلا (مثلاً 750 یا 900): "))
                price_per_methqal = float(input("🔸 قیمت هر مثقال (تومان): "))
                price_per_gram = round(price_per_methqal / METHQAL_TO_GRAM, 2)
                note = input("📝 توضیح (اختیاری): ")
                date = datetime.now().strftime("%Y-%m-%d")
                save_transaction('buy', date, weight, karat,
                                 price_per_gram, price_per_methqal, note)
                print(f"✅ خرید ثبت شد. قیمت هر گرم: {price_per_gram:,.0f} تومان")
            except ValueError:
                print("❌ ورودی نامعتبر بود.")

        elif choice == '2':
            try:
                weight = float(input("🔻 وزن فروش (گرم): "))
                karat = int(input("🔻 عیار طلا (مثلاً 750 یا 900): "))
                price_per_gram = float(input("🔻 قیمت فروش هر گرم (تومان): "))
                note = input("📝 توضیح (اختیاری): ")
                date = datetime.now().strftime("%Y-%m-%d")
                save_transaction('sell', date, weight, karat,
                                 price_per_gram, '', note)
                print("✅ فروش ثبت شد.")
            except ValueError:
                print("❌ ورودی نامعتبر بود.")

        elif choice == '3':
            show_inventory()

        elif choice == '4':
            print("👋 موفق باشید!")
            break
        else:
            print("❌ گزینه نامعتبر است.")

if __name__ == '__main__':
    main()